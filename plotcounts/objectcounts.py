'''
Script to populate an Excel spreadsheet with a report on the current object 
counts for a DataONE environment. Counts are made per object format for
each of the following types of count:

1. From CN.listObjects(count=0, formatId=formatId) to provide total number of 
   objects of that format.

2. From an unauthenticated CN.query() with query matching the formatId to 
   provide the number of publicly readable objects of that format.

3. From an unauthenticated CN.query() with query matching the formatId and
   with no obsoletedBy entry provide the number of publicly readable objects 
   of that format that have not been obsoleted.

Configuration information is read from ~/.dataone/cache.conf which is a YAML
file similar to::

  %YAML 1.2
  ---

  plotcounts:
    base_url: https://cn.dataone.org/cn
    store_file: /home/vieglais/sync/production_history/objectcounts.xlsx
    html_dest: /Users/vieglais/.dataone/plotcounts/plotcounts.html
    html_template: /Users/vieglais/.dataone/plotcounts/template/index.html
    pickle_file: /Users/vieglais/.dataone/plotcounts/objectcounts.pickle

Where:

:base_url: Is the CN base url for the environment to be examined.
:store_file: Full path to the destination XSLX file, created if not present
:html_dest: Path to HTML summary page generated (overwritten if present)
:html_template: Path of a String.Template file used to generate the html_dest
:pickle_file: Optional path of a location for a cache of downloaded content

The output is a new or updated spreadsheet that contains at least two 
worksheets.

**Summary** contains the totals for different object types (DATA, METADATA, 
RESOURCE) for the categories listed above. Column names are::
  
  A Date           Date time (UTC) when the request was run
  B Data           Number of data objects
  C Metadata       Number of metadata objects
  D Resource       Number of resource objecs
  E Total          Total number of all objects
  F pub_Data       Number of publicly accessible data objects
  G pub_Metadata   Number of publicly accessible metadata objects
  H pub_Resource   Number of publicly accessible resource objecs
  I pub_Total      Total number of publicly accessible objects
  J pubo_Data      Number of publicly accessible, not obsolete data objects
  K pubo_Metadata  Number of publicly accessible, not obsolete metadata objects
  L pubo_Resource  Number of publicly accessible, not obsolete resource objecs
  M pubo_Total     Total number of publicly accessible, not obsolete objects


A sheet named **YYYY-MM-DD** is optionally created when run with the -d or 
--detail option. That sheet contains a breakdown of counts for each object 
format.

..Note:: 

  This script is currently running on monitor.dataone.org just past 
  midnight ET, with detail sheets generated on Fridays. The folder containing 
  the spreadsheet can be accessed with the BTSync read only 
  key: BEUZJ2KNWPOKZ6ED6G4HB45AT55YW4JZ3

'''

import os
import logging
from yaml import load, Loader
from openpyxl import Workbook, load_workbook
from d1_client import cnclient, cnclient_1_1
from openpyxl.cell import get_column_letter
from datetime import datetime
from optparse import OptionParser
from string import Template
import pickle
import pprint

def getObjectFormats(base_url):
  '''
  returns a dictionary of formatIds
  '''
  client = cnclient.CoordinatingNodeClient( base_url )
  formats = client.listFormats()
  #A dictionary, keyed by formatId, each entry a list of:
  # 0: formatName
  # 1: formatType
  # 2: count by listObjects
  # 3: count by SOLR query (public access)
  # 4: count by SOLR query (public access, not obsoleted)
  res = {'formats':{},
         'timestamp': datetime.utcnow(),
         'baseurl': base_url
         }
  for format in formats.objectFormat:
    res['formats'][format.formatId] = [format.formatName, format.formatType, 0, 0, 0]
  return res


def escapeQueryTerm(term):
  '''
  + - && || ! ( ) { } [ ] ^ " ~ * ? : \
  '''
  reserved = ['+','-','&','|','!','(',')','{','}','[',']','^','"','~','*','?',':',]
  term = term.replace(u'\\',u'\\\\')
  for c in reserved:
    term = term.replace(c,u"\%s" % c)
  return term



def getHitCount(client, q):
  queryEngine = "solr"
  query='/'
  solrQuery = q
  maxrecords = 0
  fields = 'id'
  results = eval(client.query(queryEngine, query=query, 
                              q=q,
                              wt='python', 
                              fl=fields, 
                              rows=maxrecords).read())
  nHits = results['response']['numFound']
  return nHits


def countObjects(base_url, formats):
  '''
  Populates the format dictionary with the number of objects as reported by the CN
  '''
  oldclient = cnclient.CoordinatingNodeClient( base_url )
  client = cnclient_1_1.CoordinatingNodeClient( base_url )
  for formatId in formats['formats'].keys():
    res = client.listObjects(count=0, objectFormat=formatId)
    formats['formats'][formatId][2] = res.total
    q = "formatId:\"{0:s}\"".format(escapeQueryTerm(formatId))
    formats['formats'][formatId][3] = getHitCount(oldclient, q)
    q = "formatId:\"{0:s}\" AND -obsoletedBy:[* TO *]".format(escapeQueryTerm(formatId))
    formats['formats'][formatId][4] = getHitCount(oldclient, q)
    print "{0:s} : {1:d}, {2:d}, {3:d}".format(formatId, 
                                               formats['formats'][formatId][2], 
                                               formats['formats'][formatId][3],
                                               formats['formats'][formatId][4] )


def rc(row, col):
  return "{0:s}{1:d}".format(get_column_letter(col), row)


def getTotalsByType(formats, ctime=datetime.utcnow()):
  res = {'date:': ctime.strftime("%Y-%m-%d %H:%M:%S.0+00:00"),
         'total':0,
         'data':0,
         'metadata':0,
         'resource': 0,
         'total_pub':0,
         'data_pub':0,
         'metadata_pub':0,
         'resource_pub':0,
         'total_pub_no':0,
         'data_pub_no':0,
         'metadata_pub_no':0,
         'resource_pub_no':0,
         'cols':{'data':'B',
                 'metadata':'C',
                 'resource':'D',
                 'total':'E',
                 'data_pub':'F',
                 'metadata_pub':'G',
                 'resource_pub':'H',
                 'total_pub':'I',
                 'data_pub_no':'J',
                 'metadata_pub_no':'K',
                 'resource_pub_no':'L',
                 'total_pub_no':'M',
                  }
         }
  for formatId in formats.keys():
    res['total'] += formats[formatId][2]
    res['total_pub'] += formats[formatId][3]
    res['total_pub_no'] += formats[formatId][4]
    if formats[formatId][1] == 'DATA':
      res['data'] += formats[formatId][2]
      res['data_pub'] += formats[formatId][3]
      res['data_pub_no'] += formats[formatId][4]
    elif formats[formatId][1] == 'METADATA':
      res['metadata'] += formats[formatId][2]
      res['metadata_pub'] += formats[formatId][3]
      res['metadata_pub_no'] += formats[formatId][4]
    elif formats[formatId][1] == 'RESOURCE':
      res['resource'] += formats[formatId][2]
      res['resource_pub'] += formats[formatId][3]
      res['resource_pub_no'] += formats[formatId][4]
  return res


def populateWorksheet(ws, formats, label=''):
  ws.cell('A1').value = "Summary"
  ws.cell('B1').value = label
  ws.cell('B2').value = "Data"
  ws.cell('C2').value = "Metadata"
  ws.cell('D2').value = "Resource"
  ws.cell('E2').value = "Total"
  ws.cell('F2').value = "pub_Data"
  ws.cell('G2').value = "pub_Metadata"
  ws.cell('H2').value = "pub_Resource"
  ws.cell('I2').value = "pub_Total"
  ws.cell('J2').value = "pubno_Data"
  ws.cell('K2').value = "pubno_Metadata"
  ws.cell('L2').value = "pubno_Resource"
  ws.cell('M2').value = "pubno_Total"
  ws.cell('A3').value = "Total"
  
  counts = getTotalsByType(formats['formats'])
  crow = 3
  for colname in counts['cols'].keys():
    ws.cell('{0:s}{1:d}'.format(counts['cols'][colname], crow)).value = counts[colname]
  
  ws.cell('A5').value = "formatId"     #1
  ws.cell('B5').value = "type"         #2
  ws.cell('C5').value = "count"        #3
  ws.cell('D5').value = "pub_count"    #4
  ws.cell('E5').value = "pub_no_count" #5
  crow = 6
  for formatId in formats['formats'].keys():
    ws.cell(rc(crow,1)).value = formatId
    ws.cell(rc(crow,2)).value = formats['formats'][formatId][1]
    ws.cell(rc(crow,3)).value = formats['formats'][formatId][2]
    ws.cell(rc(crow,4)).value = formats['formats'][formatId][3]
    ws.cell(rc(crow,5)).value = formats['formats'][formatId][4]
    crow += 1


def updateSummarySheet(ws, ctime, formats):
  crow = ws.get_highest_row()
  if crow == 1:
    ws.cell("A1").value = "Date"
    ws.cell('B1').value = "Data"
    ws.cell('C1').value = "Metadata"
    ws.cell('D1').value = "Resource"
    ws.cell('E1').value = "Total"
    ws.cell('F1').value = "pub_Data"
    ws.cell('G1').value = "pub_Metadata"
    ws.cell('H1').value = "pub_Resource"
    ws.cell('I1').value = "pub_Total"
    ws.cell('J1').value = "pubno_Data"
    ws.cell('K1').value = "pubno_Metadata"
    ws.cell('L1').value = "pubno_Resource"
    ws.cell('M1').value = "pubno_Total"
  crow = crow + 1
  #ws.cell("A{0:d}".format(crow)).value = ctime.strftime("%Y-%m-%d")
  ws.cell("A{0:d}".format(crow)).value = ctime
  counts = getTotalsByType(formats['formats'])
  for colname in counts['cols'].keys():
    ws.cell('{0:s}{1:d}'.format(counts['cols'][colname], crow)).value = counts[colname]


def generateHtml(fnDest, formats, template_source):
  '''Generates a HTML summary of the current counts
  '''
  import locale
  locale.setlocale(locale.LC_ALL, 'en_US.utf8')
  #get counts
  counts = getTotalsByType(formats['formats'])
  #create string representations for substitution
  scounts = {}
  for k in counts.keys():
    if isinstance(counts[k], (int, long, float, complex)):
      scounts[k] = locale.format('%d', counts[k], grouping=True)
  scounts['date'] = formats['timestamp'].strftime("%Y-%m-%d %H:%M:%S.0+00:00")
  scounts['baseurl'] = formats['baseurl']
  pprint.pprint(scounts)
  res = Template(file(template_source, 'r').read())
  fout = file(fnDest,"w")
  fout.write(res.substitute(scounts))
  fout.close()


def getObjectCounts(base_url, cache_name=None, try_cache=True):
  formats = None
  if try_cache:
    if not cache_name is None:
      #attempt to load from cache
      try:
        srcfile = open(cache_name, 'rb')
        formats = pickle.load(srcfile)
        srcfile.close()
        logging.info("Object counts loaded from cache.")
        return formats
      except:
        logging.info("Failed to load from cache.")
  formats = getObjectFormats(base_url)
  countObjects(base_url, formats)
  return formats


def saveObjectCounts(formats, cache_name):
  dst = file(cache_name, 'wb')
  pickle.dump(formats, dst)
  dst.close()



if __name__ == "__main__": 
  CONFIG = os.path.join(os.environ['HOME'],".dataone/cache.conf")
  parser = OptionParser()
  parser.add_option("-l","--log",dest="loglevel",
                    help="1=DEBUG, 2=INFO, 3=WARN, 4=ERROR, 5=FATAL",
                    default=2, type="int")
  parser.add_option("-c","--config", dest="config",
                    help="Path to configuration file ({0:s})".format(CONFIG),
                    default=CONFIG)
  parser.add_option("-d","--detail", dest="doDetail",
                    help="Generate a detail set of counts (False)",
                    default=False, action="store_true")
  parser.add_option("-x","--nodownload", dest="doDownload",
                    help="Don't download new data if a cache is available (False)",
                    default=False, action="store_true")
  (options, args) = parser.parse_args()
  if options.loglevel < 1:
    options.loglevel = 1
  if options.loglevel > 5:
    options.loglevel = 5   
  logging.basicConfig(level=10*options.loglevel)

  doDetail = options.doDetail
  config = load(file(options.config, 'r'), Loader=Loader)
  store_file = os.path.join(os.environ['HOME'], ".dataone/plotcounts/plotcounts.xlsx")
  try:
    store_file = config['plotcounts']['store_file']
  except KeyError:
    pass
  pickle_file = os.path.join(os.environ['HOME'], ".dataone/plotcounts/.objectcounts.pickle")
  try:
    pickle_file = config['plotcounts']['pickle_file']
  except KeyError:
    pass
  base_url = "https://cn.dataone.org/cn"
  try:
    base_url = config['plotcounts']['base_url']
  except KeyError:
    pass
  html_dest = os.path.join(os.environ['HOME'], ".dataone/plotcounts/objectcounts.html")
  try:
    html_dest = config['plotcounts']['html_dest']
  except KeyError:
    pass
  html_template = os.path.join(os.environ['HOME'], ".dataone/plotcounts/objectcounts.template")
  try:
    html_template = config['plotcounts']['html_template']
  except KeyError:
    pass
  

  #Open the spreadsheet if available, otherwise create one.
  wb = None
  try:
    wb = load_workbook(filename = store_file)
  except:
    wb = Workbook()

  ctime = datetime.utcnow()
  formats = getObjectCounts( base_url, pickle_file, options.doDownload )
  saveObjectCounts( formats, pickle_file )
  
  summarysheet = None
  try:
    summarysheet = wb.get_sheet_by_name("Summary")
  except:
    summarysheet = create_sheet(0)
    summarysheet.title = "Summary"
  updateSummarySheet(summarysheet, ctime, formats)
  if doDetail:
    ws = wb.create_sheet()
    ws.title = ctime.strftime("%Y-%m-%d")
    label = "Sheet generated at: {0:s}".format(ctime.strftime("%Y-%m-%d %H:%M:%S.0+00:00"))
    populateWorksheet(ws, formats, label=label)
  wb.save( filename = store_file)
  generateHtml(html_dest, formats, html_template)
  print "Done."
